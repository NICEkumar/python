from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active

sheet.title = 'Language'
wb.create_sheet(title='capital')

lang = ['kan', 'telugu', 'tamil']
state = ['kan', 'telugu', 'tamil']
capital = ['kan', 'telugu', 'tamil']
code = ['kan', 'telugu', 'tamil']

sheet.cell(row=1, column=1).value = 'state'
sheet.cell(row=1, column=2).value = 'capital'
sheet.cell(1, 3).value = 'code'

